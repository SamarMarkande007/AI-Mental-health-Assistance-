"""W1-DS1 | Statistical Baseline Model
Outputs:
  - outputs/w1_ds1_statistical_baseline_model/APMS2014_MH_Regional_Model_MasterData.csv
  - outputs/w1_ds1_statistical_baseline_model/APMS2014_MH_Regional_Statistical_Model.xlsx
"""

import pandas as pd, numpy as np, os, warnings
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

DATA_DIR = "datasets"
OUT = "outputs/w1_ds1_statistical_baseline_model"
os.makedirs(OUT, exist_ok=True)

REGIONS = ["North East","North West","Yorkshire & the Humber","East Midlands",
           "West Midlands","East of England","London","South East","South West"]
REGION_POPULATION = {
    "North East":2154854,"North West":5790512,"Yorkshire & the Humber":4347165,
    "East Midlands":3782353,"West Midlands":4599064,"East of England":4875281,
    "London":6806412,"South East":7184102,"South West":4473319,
}
ENGLAND_POPULATION = sum(REGION_POPULATION.values())

def safe(v):
    try: s=str(v).strip(); return np.nan if s in ['-','nan','NaN',''] else float(s)
    except: return np.nan

def wilson_ci(prevalence_pct, n, confidence=0.95):
    if pd.isna(prevalence_pct) or pd.isna(n) or n<=0 or prevalence_pct<0:
        return np.nan, np.nan, np.nan
    p = prevalence_pct/100.0
    z = stats.norm.ppf(1-(1-confidence)/2)
    denominator = 1+z**2/n
    centre = (p+z**2/(2*n))/denominator
    margin = (z*np.sqrt(p*(1-p)/n+z**2/(4*n**2)))/denominator
    lower = max(0.0,(centre-margin)*100)
    upper = min(100.0,(centre+margin)*100)
    se = (upper-lower)/(2*z)
    return round(lower,4), round(upper,4), round(se,4)

def get_bases(df, row):
    vals = df.iloc[row, 1:10].values
    return {r: float(vals[i]) if pd.notna(vals[i]) else np.nan for i,r in enumerate(REGIONS)}

def row_to_records(df, obs_row, std_row, bases, disorder, chapter, group, table):
    obs_vals = df.iloc[obs_row, 1:10].values
    std_vals = df.iloc[std_row, 1:10].values
    recs = []
    for i,reg in enumerate(REGIONS):
        p_obs = safe(obs_vals[i]); p_std = safe(std_vals[i]); n = bases.get(reg,np.nan)
        lo,hi,se = wilson_ci(p_obs,n)
        recs.append({
            "Chapter":chapter,"Disorder_Group":group,"Disorder":disorder,"Sex":"","Region":reg,
            "Prevalence_Observed_%":p_obs,"Prevalence_AgeStd_%":p_std,"Sample_n":n,
            "CI_Lower_95_%":lo,"CI_Upper_95_%":hi,
            "CI_Width_pp":round(hi-lo,4) if not(np.isnan(hi) or np.isnan(lo)) else np.nan,
            "SE_%":se,"Source_Table":table,
        })
    return recs

all_records = []

# CH-02 CMD — sheet 2.11
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-02-tabs.xls",sheet_name="2.11",header=None)
DISORDERS_02=["Generalised anxiety disorder","Depressive episode","All phobias",
              "Obsessive compulsive disorder","Panic disorder","CMD - NOS","Any CMD"]
sex_obs_rows={"Men":6,"Women":15,"All adults":24}
sex_std_rows={"Men":36,"Women":45,"All adults":54}
base_rows_02={"Men":64,"Women":65,"All adults":66}
for sex,start in sex_obs_rows.items():
    bases=get_bases(df,base_rows_02[sex])
    for off,d in enumerate(DISORDERS_02):
        recs=row_to_records(df,start+1+off,sex_std_rows[sex]+1+off,bases,d.rstrip('b'),"02","Common Mental Disorders","2.11")
        for r in recs: r["Sex"]=sex
        all_records+=recs

# CH-04 PTSD — sheet 4.6
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-04-tabs.xls",sheet_name="4.6",header=None)
obs_rows_04={"Men":(7,8),"Women":(11,12),"All adults":(15,16)}
std_rows_04={"Men":(20,21),"Women":(24,25),"All adults":(28,29)}
base_rows_04={"Men":32,"Women":33,"All adults":34}
for sex,(r1,r2) in obs_rows_04.items():
    bases=get_bases(df,base_rows_04[sex])
    for ind,ro,rs in [("Trauma experienced",r1,std_rows_04[sex][0]),("PTSD screen positive",r2,std_rows_04[sex][1])]:
        recs=row_to_records(df,ro,rs,bases,ind,"04","PTSD","4.6")
        for r in recs: r["Sex"]=sex
        all_records+=recs

# CH-05 Psychosis — sheet 5.5
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-05-tabs.xls",sheet_name="5.5",header=None)
for sex,row,br in [("Men",5,10),("Women",6,11),("All adults",7,12)]:
    bases=get_bases(df,br)
    recs=row_to_records(df,row,row,bases,"Psychotic disorder (past year)","05","Psychotic Disorder","5.5")
    for r in recs: r["Sex"]=sex
    all_records+=recs

# CH-07 Personality Disorder — sheet 7.14
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-07-tabs.xls",sheet_name="7.14",header=None)
for sex,obs,std,br in [("Men",6,11,16),("Women",7,12,17),("All adults",8,13,18)]:
    bases=get_bases(df,br)
    recs=row_to_records(df,obs,std,bases,"Any personality disorder (SAPAS)","07","Personality Disorder","7.14")
    for r in recs: r["Sex"]=sex
    all_records+=recs

# CH-08 ADHD — sheet 8.6
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-08-tabs.xls",sheet_name="8.6",header=None)
obs_rows_08={"Men":(7,8),"Women":(11,12),"All adults":(15,16)}
std_rows_08={"Men":(20,21),"Women":(24,25),"All adults":(28,29)}
base_rows_08={"Men":32,"Women":33,"All adults":34}
for sex,(r1,r2) in obs_rows_08.items():
    bases=get_bases(df,base_rows_08[sex])
    for ind,ro,rs in [("ADHD screen (4+ symptoms)",r1,std_rows_08[sex][0]),("ADHD screen (all 6 symptoms)",r2,std_rows_08[sex][1])]:
        recs=row_to_records(df,ro,rs,bases,ind,"08","ADHD","8.6")
        for r in recs: r["Sex"]=sex
        all_records+=recs

# CH-09 Bipolar — sheet 9.4
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-09-tabs.xls",sheet_name="9.4",header=None)
for sex,obs,std,br in [("Men",6,11,16),("Women",7,12,17),("All adults",8,13,18)]:
    bases=get_bases(df,br)
    recs=row_to_records(df,obs,std,bases,"Bipolar disorder screen positive","09","Bipolar Disorder","9.4")
    for r in recs: r["Sex"]=sex
    all_records+=recs

# CH-10 Alcohol — sheet 10.6, ch02 bases as proxy
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-10-tabs.xls",sheet_name="10.6",header=None)
df_b=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-02-tabs.xls",sheet_name="2.11",header=None)
inds_10=["Non-drinker/low risk (AUDIT 0-7)","Hazardous drinking (AUDIT 8-15)",
         "Harmful/mild dependence (AUDIT 16-19)","Probable dependence (AUDIT 20+)",
         "Harmful/dependent drinking (AUDIT 8+)","High-risk drinking (AUDIT 16+)"]
offsets=[0,1,2,3,5,6]
for sex,obs_s,br in [("Men",6,64),("Women",15,65),("All adults",24,66)]:
    bases=get_bases(df_b,br)
    for off,ind in zip(offsets,inds_10):
        recs=row_to_records(df,obs_s+1+off,obs_s+1+off,bases,ind,"10","Alcohol Use","10.6")
        for r in recs: r["Sex"]=sex
        all_records+=recs

# CH-11 Drug — sheet 11.11
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-11-tabs.xls",sheet_name="11.11",header=None)
sex_rows_11={"Men":(6,7,8),"Women":(11,12,13),"All adults":(16,17,18)}
base_rows_11={"Men":21,"Women":22,"All adults":23}
for sex,rows in sex_rows_11.items():
    bases=get_bases(df,base_rows_11[sex])
    for ind,r in zip(["Cannabis only dependence","Other drug(s) dependence","Any drug dependence"],rows):
        recs=row_to_records(df,r,r,bases,ind,"11","Drug Dependence","11.11")
        for rec in recs: rec["Sex"]=sex
        all_records+=recs

# CH-12 Suicide — sheet 12.7
df=pd.read_excel(f"{DATA_DIR}/apms-2014-ch-12-tabs.xls",sheet_name="12.7",header=None)
sex_rows_12={"Men":(7,8,9),"Women":(11,12,13),"All adults":(15,16,17)}
std_rows_12={"Men":(21,22,23),"Women":(25,26,27),"All adults":(29,30,31)}
base_rows_12={"Men":34,"Women":35,"All adults":36}
for sex,rows in sex_rows_12.items():
    bases=get_bases(df,base_rows_12[sex])
    for ind,ro,rs in zip(["Lifetime suicidal thoughts","Lifetime suicide attempts","Lifetime self-harm"],rows,std_rows_12[sex]):
        recs=row_to_records(df,ro,rs,bases,ind,"12","Suicidal Behaviour","12.7")
        for r in recs: r["Sex"]=sex
        all_records+=recs

# Build master dataframe 
master_df = pd.DataFrame(all_records)
master_df["Region_Population"] = master_df["Region"].map(REGION_POPULATION)
master_df["England_Population"] = ENGLAND_POPULATION
master_df["Population_Weight"]  = master_df["Region_Population"] / ENGLAND_POPULATION
col_order = ["Chapter","Disorder_Group","Disorder","Sex","Region",
             "Prevalence_Observed_%","Prevalence_AgeStd_%","Sample_n",
             "CI_Lower_95_%","CI_Upper_95_%","CI_Width_pp","SE_%",
             "Region_Population","Population_Weight","Source_Table"]
master_df = master_df[col_order]

all_adults = master_df[master_df["Sex"]=="All adults"].copy()

# Descriptive statistics
stats_df = (
    all_adults.groupby(["Disorder_Group","Disorder"])["Prevalence_Observed_%"]
    .agg(N_regions="count",Mean="mean",Median="median",Std_Dev="std",
         Min="min",Max="max",Range=lambda x:x.max()-x.min(),
         CV_pct=lambda x:(x.std()/x.mean()*100) if x.mean()!=0 else np.nan)
    .round(3).reset_index()
)

# Composite burden score 
key_combos = [
    ("Common Mental Disorders","Any CMD"),
    ("PTSD","PTSD screen positive"),
    ("Psychotic Disorder","Psychotic disorder (past year)"),
    ("Personality Disorder","Any personality disorder (SAPAS)"),
    ("ADHD","ADHD screen (4+ symptoms)"),
    ("Bipolar Disorder","Bipolar disorder screen positive"),
    ("Alcohol Use","High-risk drinking (AUDIT 16+)"),
    ("Drug Dependence","Any drug dependence"),
    ("Suicidal Behaviour","Lifetime suicidal thoughts"),
]
sub_all = all_adults[all_adults.apply(lambda r:(r["Disorder_Group"],r["Disorder"]) in key_combos,axis=1)].copy()
matrix = sub_all.pivot_table(index="Region",columns="Disorder",values="Prevalence_Observed_%")
z_matrix = (matrix-matrix.mean())/matrix.std()
composite = z_matrix.mean(axis=1).sort_values(ascending=False).round(4)
composite_df = composite.reset_index()
composite_df.columns = ["Region","Composite_Z_Score"]
composite_df["Burden_Category"] = composite_df["Composite_Z_Score"].apply(
    lambda z: "High" if z>0.3 else ("Above average" if z>0 else ("Below average" if z>-0.3 else "Low")))
composite_df["Region_Population"] = composite_df["Region"].map(REGION_POPULATION)

# England national summary 
summary_rows = []
for grp,dis in key_combos:
    sub = all_adults[(all_adults["Disorder_Group"]==grp)&(all_adults["Disorder"]==dis)].dropna(subset=["Prevalence_Observed_%"])
    if sub.empty: continue
    eng_avg = (sub["Prevalence_Observed_%"]*sub["Population_Weight"]).sum()/sub["Population_Weight"].sum()
    eng_n = int(sub["Sample_n"].sum())
    lo,hi,se = wilson_ci(eng_avg,eng_n)
    min_reg = sub.loc[sub["Prevalence_Observed_%"].idxmin(),"Region"]
    max_reg = sub.loc[sub["Prevalence_Observed_%"].idxmax(),"Region"]
    summary_rows.append({
        "Disorder_Group":grp,"Disorder":dis,
        "England_Avg_%":round(eng_avg,3),"Total_n":eng_n,
        "CI_Lower_%":lo,"CI_Upper_%":hi,"CI_Width_pp":round(hi-lo,3),
        "Lowest_Region":f"{min_reg} ({sub['Prevalence_Observed_%'].min():.2f}%)",
        "Highest_Region":f"{max_reg} ({sub['Prevalence_Observed_%'].max():.2f}%)",
        "Range_pp":round(sub["Prevalence_Observed_%"].max()-sub["Prevalence_Observed_%"].min(),3),
    })
summary_df = pd.DataFrame(summary_rows)

# Plot 
fig, axes = plt.subplots(1,2,figsize=(16,12))
fig.suptitle("APMS 2014 -- Mental Health Prevalence by Region",fontsize=15,fontweight="bold",y=1.01)


ax1=axes[0]
comp_sorted=composite_df.sort_values("Composite_Z_Score",ascending=True)
colors2=["#d62728" if z>0.3 else "#FFC000" if z>0 else "#92D050" for z in comp_sorted["Composite_Z_Score"]]
bars2=ax1.barh(comp_sorted["Region"],comp_sorted["Composite_Z_Score"],color=colors2,alpha=0.85,height=0.6)
ax1.axvline(0,color="black",linewidth=1)
for bar,val in zip(bars2,comp_sorted["Composite_Z_Score"]):
    sign="+" if val>=0 else ""
    ax1.text(val+0.01 if val>=0 else val-0.01,bar.get_y()+bar.get_height()/2,
             sign+str(round(val,3)),va="center",ha="left" if val>=0 else "right",fontsize=9)
ax1.set_xlabel("Composite Z-Score (0 = England average)")
ax1.set_title("Composite MH Burden Score\n(9 disorders combined)",fontweight="bold")
patches=[mpatches.Patch(color="#d62728",label="High burden (z>0.3)"),
         mpatches.Patch(color="#FFC000",label="Above average"),
         mpatches.Patch(color="#92D050",label="Below average")]
ax1.legend(handles=patches,fontsize=8)

ax2=axes[1]
heat_disorders=["Any CMD","PTSD screen positive","Psychotic disorder (past year)",
                "Any personality disorder (SAPAS)","ADHD screen (4+ symptoms)",
                "Bipolar disorder screen positive","High-risk drinking (AUDIT 16+)",
                "Any drug dependence","Lifetime suicidal thoughts"]
heat_data=all_adults[all_adults["Disorder"].isin(heat_disorders)].pivot_table(index="Disorder",columns="Region",values="Prevalence_Observed_%")
heat_data.index=[d[:30] for d in heat_data.index]
heat_norm=heat_data.apply(lambda row:(row-row.min())/(row.max()-row.min()) if row.max()>row.min() else row,axis=1)
region_short=[r.replace("Yorkshire & the Humber","Yorks.") for r in REGIONS]
sns.heatmap(heat_norm,ax=ax2,cmap="RdYlGn_r",annot=heat_data.round(1),fmt=".1f",
            annot_kws={"size":7},linewidths=0.5,cbar=False,xticklabels=region_short)
ax2.set_title("Prevalence Heatmap (%) -- All Adults\n(row-normalised: green=low, red=high)",fontweight="bold")
ax2.tick_params(axis="x",rotation=40,labelsize=7); ax2.tick_params(axis="y",rotation=0,labelsize=7)
ax2.set_xlabel(""); ax2.set_ylabel("")

plt.tight_layout()
plt.savefig(f"{OUT}/APMS2014_regional_prevalence.png",dpi=140,bbox_inches="tight")
plt.close()

# Export CSV
master_df.to_csv(f"{OUT}/APMS2014_MH_Regional_Model_MasterData.csv",index=False)

# Export Excel 
def hdr(bold=True,color="FFFFFFFF",size=10):
    return Font(name="Calibri",bold=bold,color=color,size=size)
def fill(hex_c):
    return PatternFill("solid",fgColor=hex_c)
def thin_border():
    s=Side(style="thin",color="FFBFBFBF")
    return Border(left=s,right=s,top=s,bottom=s)
def ctr():
    return Alignment(horizontal="center",vertical="center",wrap_text=True)

DARK="FF1F3864"; MED="FF2E5FA3"

xlsx_path=f"{OUT}/APMS2014_MH_Regional_Statistical_Model.xlsx"
with pd.ExcelWriter(xlsx_path,engine="openpyxl") as writer:

    master_df.to_excel(writer,sheet_name="All Data",index=False)
    ws=writer.sheets["All Data"]
    for ci in range(1,len(master_df.columns)+1):
        c=ws.cell(row=1,column=ci); c.font=hdr(); c.fill=fill(DARK); c.alignment=ctr(); c.border=thin_border()
    ws.auto_filter.ref=ws.dimensions; ws.freeze_panes="A2"

    summary_df.to_excel(writer,sheet_name="England Summary",index=False)
    ws2=writer.sheets["England Summary"]
    for ci in range(1,len(summary_df.columns)+1):
        c=ws2.cell(row=1,column=ci); c.font=hdr(); c.fill=fill(DARK); c.alignment=ctr(); c.border=thin_border()
    for row in range(2,len(summary_df)+2):
        for ci in range(1,len(summary_df.columns)+1):
            c=ws2.cell(row=row,column=ci); c.border=thin_border()
            c.fill=fill("FFF2F6FC") if row%2==0 else fill("FFFFFFFF")
    ws2.freeze_panes="A2"

    composite_df.to_excel(writer,sheet_name="Composite Burden Score",index=False)
    ws3=writer.sheets["Composite Burden Score"]
    for ci in range(1,len(composite_df.columns)+1):
        c=ws3.cell(row=1,column=ci); c.font=hdr(); c.fill=fill(DARK); c.alignment=ctr(); c.border=thin_border()
    for row in range(2,len(composite_df)+2):
        for ci in range(1,len(composite_df.columns)+1):
            ws3.cell(row=row,column=ci).border=thin_border()
    ws3.freeze_panes="A2"

    any_cmd_full=all_adults[all_adults["Disorder"]=="Any CMD"][
        ["Region","Prevalence_Observed_%","Prevalence_AgeStd_%","Sample_n",
         "CI_Lower_95_%","CI_Upper_95_%","CI_Width_pp","SE_%","Population_Weight"]
    ].copy().reset_index(drop=True)
    any_cmd_full.to_excel(writer,sheet_name="CMD by Region (CI)",index=False)
    ws4=writer.sheets["CMD by Region (CI)"]
    for ci in range(1,len(any_cmd_full.columns)+1):
        c=ws4.cell(row=1,column=ci); c.font=hdr(); c.fill=fill(MED); c.alignment=ctr(); c.border=thin_border()
    ws4.auto_filter.ref=ws4.dimensions

    stats_df.to_excel(writer,sheet_name="Descriptive Statistics",index=False)
    ws5=writer.sheets["Descriptive Statistics"]
    for ci in range(1,len(stats_df.columns)+1):
        c=ws5.cell(row=1,column=ci); c.font=hdr(); c.fill=fill(DARK); c.alignment=ctr(); c.border=thin_border()
    ws5.auto_filter.ref=ws5.dimensions

print(f"[DS1] Done — {len(master_df):,} rows")
print(f" APMS2014_MH_Regional_Model_MasterData.csv")
print(f" APMS2014_MH_Regional_Statistical_Model.xlsx  (5 sheets)")
print(f" APMS2014_regional_prevalence.png")